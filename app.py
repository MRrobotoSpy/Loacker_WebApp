import streamlit as st
import cv2
from PIL import Image
import numpy as np
import json 
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.colors import ListedColormap
import matplotlib.colors

import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font, Border, Side




def create_excel_file(cluster_images, fat_coverage, choco_coverage,highlighted_image, original,coverage_percentages,suggested_area,  intensity):
  
  
    file_path = 'analysis_results.xlsx'

    # Create a DataFrame to store the data
    data = {'Fat Percentage': [fat_coverage], 'Choco Percentage': [choco_coverage]}



   

    for i in range(len(coverage_percentages)): 
        if i == 1:
            data[f'Choccolate'] = [f'{round(coverage_percentages[i],2)}%']
        elif i != 0:
            data[f'Intensity_{i-1}'] = [f'{round(coverage_percentages[i],2)}%']
       
        

    df = pd.DataFrame(data)

    # Write the DataFrame to an Excel file
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Sheet1', index=False, startrow=6)
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        # Set column width for the first 6 columns
        for col_idx in range(1, 15):
            worksheet.column_dimensions[get_column_letter(col_idx)].width = 20  # Adjust the width as needed

        worksheet.column_dimensions[get_column_letter(3)].width = 30  # Adjust the width as needed
        worksheet.column_dimensions[get_column_letter(13)].width = 40  # Adjust the width as needed

        # Find the last written column
        last_column = max(worksheet.max_column, 1)


        worksheet.cell(row=40, column=1 ,value=f'Suggested Intensity')

        worksheet.cell(row=40, column=2, value=f'Suggested Area')


        worksheet.cell(row=41, column=2 ,value=suggested_area)
        worksheet.cell(row=41, column=1, value=intensity)


        img = Image.fromarray(np.array(original))
        img_path = 'original.png'
        img.save(img_path)

        # Convert the saved image to BytesIO
        with open(img_path, "rb") as img_file:
            image_bytes = io.BytesIO(img_file.read())

        # Reset the BytesIO object to the beginning
        image_bytes.seek(0)

        # Add the image to the Excel worksheet with specified dimensions and position
        img_excel = ExcelImage(image_bytes)
        img_excel.width = 450  # Set the width of the image
        img_excel.height = 450  # Set the height of the image
        worksheet.add_image(img_excel, f'{get_column_letter(5)}15')
        worksheet.cell(row=14, column=5, value=f'Original')



        # Assuming highlighted_image is a BytesIO object
        highlighted_image.seek(0)  # Make sure to reset the position to the beginning

        # Add the image to the Excel worksheet with specified dimensions and position
        img_excel = ExcelImage(highlighted_image)
        img_excel.width = 450  # Set the width of the image
        img_excel.height = 450  # Set the height of the image
        worksheet.add_image(img_excel, f'{get_column_letter(8)}15')
        worksheet.cell(row=14, column=8, value=f'Predicted Area')

        # Add remaining images to Excel worksheet with a distance of 3 columns
        for i, img_array in enumerate(cluster_images[1:], start=2):
            # Convert the NumPy array to an Image
            img = Image.fromarray(img_array)

            # Save the image to a file
            img_path = f'image{i}.png'
            img.save(img_path)

            # Convert the saved image to BytesIO
            with open(img_path, "rb") as img_file:
                img_bytes = io.BytesIO(img_file.read())

            # Reset the BytesIO object to the beginning
            img_bytes.seek(0)

            # Add the image to the Excel worksheet with specified dimensions and position
            img_excel = ExcelImage(img_bytes)
            img_excel.width = 100  # Set the width of the image
            img_excel.height = 100  # Set the height of the image
            worksheet.add_image(img_excel, f'{get_column_letter(last_column + 5)}{16+ i * 4}')

            # Add intensity under each image in the last possible column
            if(i-2 == 0):
                worksheet.cell(row=18 + i * 4, column=last_column + 4, value=f'Choccolate')
            else:
                worksheet.cell(row=18 + i * 4, column=last_column + 4, value=f'{i-2}')
            worksheet.cell(row=15, column=last_column + 4, value=f'Intensity Detected')
            worksheet.cell(row=15, column=last_column + 5, value=f'Segmentation Area')

                      
                # Ignore the first element and find non-zero indices and corresponding non-zero coverages
        non_zero_indices = [i for i, percentage in enumerate(coverage_percentages[1:]) if percentage > 0]
        non_zero_coverage = [coverage_percentages[i + 1] for i in non_zero_indices]

        # Custom labels for non-zero values
        custom_labels = [f'Intensity {i-1}' for i in range(1, len(non_zero_coverage) + 1)]

        custom_labels[0] = "Choccolate"

        # Add a pie chart using non-zero coverage_percentages
        pie_chart_path = 'pie_chart.png'
        plt.figure(figsize=(5, 5))
        plt.pie(non_zero_coverage, labels=custom_labels, autopct='%1.1f%%')
        plt.title('Coverage Percentages')
        plt.savefig(pie_chart_path)
        plt.close()


        # Add the pie chart to Excel worksheet
        img_excel = ExcelImage(pie_chart_path)
        img_excel.width = 450  # Set the width of the image
        img_excel.height = 450  # Set the height of the image
        worksheet.add_image(img_excel, f'{get_column_letter(2)}15')
        worksheet.cell(row=14, column=2, value=f'Pie Chart')

         # Add the list of columns at a specific row (e.g., row 0)
        column_headers = [
            'Product', 'Year', 'F&Q code or material number', 'Format (g)', 'Packing material (code)',
            'Productionsdate', 'Aimed SL (shelf life)', 'Aim of the test', 'Project Nr', 'Identifier',
            'Comments', 'Storage conditions', 'Intended/proposed sensory evaluation', 'Product variants'
        ]

        for idx, header in enumerate(column_headers):
            cell = worksheet.cell(row=1, column= 1 + idx, value=header)

            # Apply bold font to text
            cell.font = Font(bold=True)

            # Apply bold border to the cell
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            cell.border = border



    return file_path








def map_to_percentage(num):
    if num <= 2:
        return 0
    elif num < 25:
        return 1
    elif num < 50:
        return 2
    elif num < 75:
        return 3
    else:
        return 4





def get_cluster_data_by_name(json_file, name):
    # Load the JSON file
    with open(json_file, 'r') as file:
        data = json.load(file)

    # Retrieve data based on the name
    if name in data:
        return np.array(data[name])
    else:
        print(f"Cluster data for {name} not found.")
        return None


def remove_background(image):
    """
     # Convert to grayscale
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

    # Threshold 160 and set the rest to 255 (remove the border)
    ret, thresh = cv2.threshold(gray, 190, 255, cv2.THRESH_BINARY)

    # Set all values equal to 255 to 0 (black)
    image[thresh == 255] = 0

    kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (5, 5))
    erosion = cv2.erode(image, kernel, iterations=1)
    
    return erosion  # Return the processed image
    """
    image_rgb = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)

    lower_blue = np.array([155, 155, 150])
    upper_blue = np.array([255, 255, 255])
    
    # Create the mask
    mask = cv2.inRange(image_rgb, lower_blue, upper_blue)
    
    # Apply the mask to the image
    result = cv2.bitwise_and(image_rgb, image_rgb, mask=cv2.bitwise_not(mask))
    result_inverted_rgb = cv2.cvtColor(result, cv2.COLOR_BGR2RGB)
    return result_inverted_rgb
                            




def create_cluster_images(image, cluster_assignment):
    # Get unique cluster labels
    unique_clusters = np.unique(cluster_assignment)

    # Create an empty list to store cluster images
    cluster_images = []

    # Define the number of clusters you always expect (in your case, 6)
    num_clusters = 6

    for i in range(num_clusters):
        # Check if the current cluster label exists in the unique_clusters
        if i in unique_clusters:
            # Extract pixels corresponding to the current cluster
            cluster_pixels = image[cluster_assignment == i]

            # Reshape the pixels back to the image shape
            cluster_image = np.zeros_like(image)
            cluster_image[cluster_assignment == i] = cluster_pixels

            # Append the cluster image to the list
            cluster_images.append(cluster_image)
        else:
            # If the cluster doesn't exist, append a black image
            cluster_images.append(np.zeros_like(image))

    return cluster_images

# Modify analyze_clusters to include the visualization part
def analyze_clusters(image, chocolate_type):
    h, w, _ = image.shape
    pixels = image.reshape((h * w, 3))

    # Assign each pixel to the closest cluster in chocolate_type
    cluster_assignment = np.argmin(np.linalg.norm(pixels[:, np.newaxis, :] - chocolate_type, axis=2), axis=1)

    # Reshape the labels back to the image shape
    cluster_assignment = cluster_assignment.reshape((h, w))

    
    # Create images for each cluster
    cluster_images = create_cluster_images(image, cluster_assignment)



    custom_cmap = matplotlib.colors.LinearSegmentedColormap.from_list("", ["black", "gray", "green", "yellow", "orange", "red"])


     # Plot the image
    plt.imshow(cluster_assignment, cmap=custom_cmap)
    plt.title('Area')

    # Remove x and y axis scales
    plt.axis('off')
    cbar = plt.colorbar(label='Intensity Index', cmap=custom_cmap)

    # Set custom tick labels for the colorbar
    cbar.ax.set_yticklabels(['Background','Choccolate', 'Intensity 1', 'Intensity 2', 'Intensity 3', 'Intensity 4'])  # Modify this line based on your requirements

 


    # Save the plot as an image in a variable
    image_bytes = io.BytesIO()
    plt.savefig(image_bytes, format='png')
    image_bytes.seek(0)

    return image_bytes, cluster_images




def cluster_occupation(original_image, cluster_images):
   

     # Check if each pixel is not black ([0, 0, 0])
    non_black_pixels =  np.all(original_image != [0, 0, 0], axis=-1)

    # Count the number of non-black pixels
    total_pixels = np.sum(non_black_pixels)



    # Initialize a list to store occupation percentages for each cluster
    occupation_percentages = []

 

    # Calculate the occupation percentage for each cluster
    for cluster_image in cluster_images:

        # Extract pixels corresponding to the current cluster from non-background pixels
        non_black_cluster_pixels =  np.all(cluster_image != [0, 0, 0], axis=-1)
        cluster_pixels =  np.sum(non_black_cluster_pixels   )



        occupation_percentage = (cluster_pixels / total_pixels) * 100.0
        occupation_percentages.append(occupation_percentage)

    return occupation_percentages



def rgb_analysis(image_input,key):

    
    image = remove_background(image_input)

    chocolate_type = get_cluster_data_by_name("cluster_centers.json", key)

    area , cluster_images = analyze_clusters(image,chocolate_type)

    occupation_percentages = cluster_occupation(image, cluster_images)



    return   area , cluster_images,occupation_percentages






from tensorflow.keras.models import load_model
from tensorflow.keras.preprocessing.image import img_to_array
from tensorflow.keras.applications.imagenet_utils import preprocess_input

# Load the model
model_path = 'classifier_loacker.h5'
model = load_model(model_path)

def main():
    st.title("Image Analysis App")

    uploaded_file = st.file_uploader("Choose an image...", type=["jpg", "png"])

    keys = [
        "Chocolate_bar - Dark",
        "Chocolate_bar - Milk", 
        "Gardena - Dark",
        "Gardena - Milk",
        "Gardena_Tortina - Dark",
        "Gardena_Tortina - Milk",
        "Gardena_fingers - Dark",
        "Gardena_fingers - Milk",
        "Loacker_and_Cereals - Milk",
        "Napolitaner_Mini - Milk",
        "Patisserie_Cappuccino - Milk",
        "Patisserie_Coconut - Milk",
        "Patisserie_Crème_Noisette - Milk",
        "Patisserie_Noir - Dark", 
        "Patisserie_Orange - Dark", 
    ]

    class_name_to_index = [
        "Chocolate_bar_Dark" ,
        "Chocolate_bar_Milk",
        "Gardena_Dark",
        "Gardena_Milk",
        "Gardena_Tortina_Dark",
        "Gardena_Tortina_Milk",
        "Gardena_fingers_Dark",
        "Gardena_fingers_Milk",
        "Loacker_and_Cereals_Milk",
        "Napolitaner_Mini_Milk",
        "Patisserie_Cappuccino_Milk",
        "Patisserie_Coconut_Milk",
        "Patisserie_Crème_Noisette_Milk",
        "Patisserie_Noir_Dark",
        "Patisserie_Orange_Dark"
    ]



    


    # Streamlit selectbox
    selected_key = st.selectbox("Select a key", keys)


    if uploaded_file is not None:
        # Convert the uploaded image to a format compatible with OpenCV
        image = Image.open(uploaded_file)
        image2 = Image.open(uploaded_file)


        image_np = np.array(image)


        st.image(image, caption="Uploaded Image", use_column_width=True)


        image2 = image2.resize((150, 150))  # Resize to match the input size of your model
        image_array = img_to_array(image2)
        image_array = np.expand_dims(image_array, axis=0)

        # Perform prediction using the model
        prediction = model.predict(image_array)
        predicted_class_index = np.argmax(prediction, axis=1)[0]
        predicted_class_label = keys[predicted_class_index]
        st.subheader(f"Predicted Class: {predicted_class_label}")
        st.subheader(f"Predicted index: {predicted_class_index}")

       




        # Confirm button
        if st.button("Perform Image Analysis", key="centered_button"):
            st.header("Analysis Results")

            # Perform image analysis
            area_image , cluster_images, occupation_percentages=  rgb_analysis(image_np, selected_key)

              # Display the "Area" image
            st.image(area_image, caption="Area", use_column_width=True)



            print(type(area_image))
            print(type(image))



            # Display the results using Streamlit components
            col1, col2, col3 = st.columns(3)        

            # Display images in the first column
            col1.image(cluster_images[1], caption=f'Choccolate', use_column_width=True)
            col2.image(cluster_images[2], caption=f'Intensity 1', use_column_width=True)

            # Display images in the second column
            col3.image(cluster_images[3], caption=f'Intensity 2', use_column_width=True)
            col1.image(cluster_images[4], caption=f'Intensity 3', use_column_width=True)

            # Display images in the second row
            col2.image(cluster_images[5], caption=f'Intensity 4', use_column_width=True)



            suggested_area = map_to_percentage(100-occupation_percentages[1])

            # Reverse the array and find the index of the first non-zero element
            reversed_arr = np.flip(occupation_percentages)
            last_nonzero_index = np.argmax(reversed_arr != 0)

            # Return the last non-zero index (converted to the original index)
            intensity =  len(reversed_arr) - 2 - last_nonzero_index



            choco_coverage = round(occupation_percentages[1],2)
            fat_coverage =  round(100 - choco_coverage , 2)


            st.header("Suggestions")
            st.text(f'Choccolate Coverage: {choco_coverage}%')
            st.text(f'Total Fat bloom coverage: {fat_coverage}%')
            st.text(f'Suggested Area: {suggested_area}')
            
            
            if(intensity != 0):
             st.text(f'Suggested intensity: {intensity}\n\n')
            else:
             st.text('All Chocolate\n\n')



            st.header("Intensity Results")

            for i in range(len(occupation_percentages)): 
                if occupation_percentages[i] != 0 and i > 1:
                    st.text(f'Intensity: {i-1}, Coverage: {round(occupation_percentages[i],2)}%')


            
            # Generate and display a downloadable Excel file
            excel_file = create_excel_file(cluster_images, fat_coverage, choco_coverage,area_image, image,occupation_percentages, suggested_area,intensity)

            st.download_button(
                label="Download Excel File",
                data=open(excel_file, 'rb').read(),
                file_name="analysis_results.xlsx",
                key="download_button"
            )









if __name__ == "__main__":
    main()